#!/usr/bin/env python3
"""
Excel Browser Rendering Payload Database Exporter
=================================================

This script exports Excel browser rendering payload database to Excel format,
focusing on Excel files opened and rendered in web browsers.

Author: SNGWN
Legal Notice: For authorized security testing only.
"""

import json
import pandas as pd
import os
import sys
import logging
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

VERSION = "2.0"

# Configure logging
def setup_logging(log_level=logging.INFO, log_file=None):
    """Setup logging configuration with optional file output"""
    handlers = [logging.StreamHandler(sys.stdout)]
    
    if log_file:
        handlers.append(logging.FileHandler(log_file, encoding='utf-8'))
    
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )
    
    return logging.getLogger(__name__)

logger = setup_logging()

def find_latest_payload_database():
    """Find Excel browser payload database file"""
    logger.info("Searching for Excel browser payload database")
    
    if os.path.exists('excel_payloads.json'):
        logger.info("Found excel_payloads.json")
        return 'excel_payloads.json'
    
    logger.error("Excel payload database not found (excel_payloads.json)")
    return None

def load_payload_database(file_path):
    """Load and validate Excel browser payload database"""
    logger.info(f"Loading Excel browser payload database from {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        payloads = data.get('payloads', [])
        metadata = data.get('metadata', {})
        
        logger.info(f"Loaded {len(payloads)} Excel browser payloads")
        logger.info(f"Database focus: {metadata.get('focus', 'Excel browser rendering')}")
        logger.info(f"Target formats: {', '.join(metadata.get('target_formats', []))}")
        logger.info(f"Browser targets: {', '.join(metadata.get('browser_targets', []))}")
        
        return data
        
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {file_path}: {e}")
        return None
    except IOError as e:
        logger.error(f"Cannot read file {file_path}: {e}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error loading {file_path}: {e}")
        return None

def create_excel_workbook(data):
    """Create comprehensive Excel workbook with multiple sheets"""
    logger.info("Creating Excel workbook")
    
    payloads = data.get('payloads', [])
    metadata = data.get('metadata', {})
    
    # Create main DataFrame from payloads
    df = pd.DataFrame(payloads)
    logger.info(f"Processing {len(df)} payloads for Excel export")
    
    # Create workbook with multiple sheets
    wb = Workbook()
    wb.remove(wb.active)
    
    # 1. Main payloads sheet
    ws_payloads = wb.create_sheet("Payloads", 0)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_payloads.cell(row=r_idx, column=c_idx)
            cell.value = value
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # 2. Metadata sheet
    ws_metadata = wb.create_sheet("Metadata", 1)
    ws_metadata['A1'] = "Key"
    ws_metadata['B1'] = "Value"
    ws_metadata['A1'].font = Font(bold=True)
    ws_metadata['B1'].font = Font(bold=True)
    
    row_idx = 2
    for key, value in metadata.items():
        ws_metadata[f'A{row_idx}'] = key
        ws_metadata[f'B{row_idx}'] = str(value)
        row_idx += 1
    
    # 3. Statistics sheet
    ws_stats = wb.create_sheet("Statistics", 2)
    ws_stats['A1'] = "Category"
    ws_stats['B1'] = "Count"
    
    categories = {}
    for payload in payloads:
        category = payload.get('category', 'Unknown')
        categories[category] = categories.get(category, 0) + 1
    
    row_idx = 2
    for category, count in sorted(categories.items()):
        ws_stats[f'A{row_idx}'] = category
        ws_stats[f'B{row_idx}'] = count
        row_idx += 1
    
    logger.info(f"Created workbook with {len(wb.sheetnames)} sheets")
    return wb

def save_workbook(wb, output_dir=None):
    """Save workbook to file"""
    if output_dir is None:
        output_dir = 'output'
    
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"excel_browser_payload_database_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    try:
        wb.save(filepath)
        logger.info(f"Workbook saved to {filepath}")
        return filepath
    except IOError as e:
        logger.error(f"Failed to save workbook: {e}")
        return None

def main():
    """Main function"""
    parser = argparse.ArgumentParser(
        description="Export Excel browser rendering payload database to Excel workbook"
    )
    parser.add_argument(
        '--log-level',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help='Logging level (default: INFO)'
    )
    parser.add_argument(
        '--log-file',
        help='Optional log file path'
    )
    parser.add_argument(
        '--output-dir',
        default='output',
        help='Output directory for Excel workbook (default: output/)'
    )
    
    args = parser.parse_args()
    
    # Reconfigure logging with user args
    global logger
    log_level = getattr(logging, args.log_level)
    logger = setup_logging(log_level, args.log_file)
    
    logger.info(f"Excel Browser Payload Database Exporter v{VERSION}")
    logger.info("=" * 50)
    
    # Find payload database
    db_file = find_latest_payload_database()
    if not db_file:
        logger.error("Cannot proceed without payload database")
        return False
    
    # Load data
    data = load_payload_database(db_file)
    if not data:
        logger.error("Failed to load payload database")
        return False
    
    # Create workbook
    wb = create_excel_workbook(data)
    if not wb:
        logger.error("Failed to create workbook")
        return False
    
    # Save workbook
    filepath = save_workbook(wb, args.output_dir)
    if not filepath:
        logger.error("Failed to save workbook")
        return False
    
    logger.info("Export completed successfully")
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
