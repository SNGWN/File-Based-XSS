#!/usr/bin/env python3
"""
XSS-PDF Payload Database Excel Exporter
========================================

This script exports the XSS-PDF payload database to Excel format with the same
objective and research data level as the original PDF generation tools.

SAME OBJECTIVES:
- Security testing and authorized penetration testing
- Comprehensive payload database for PDF sandbox escape research
- Browser-specific targeting (Chrome PDFium, Firefox PDF.js, Safari PDFKit, Adobe Reader, Edge)
- Payload categorization (DOM access, file system, command execution, sandbox escape, network exfiltration)

SAME RESEARCH DATA LEVEL:
- 50+ CVE references across all PDF rendering libraries
- Academic papers on PDF security and sandbox escapes
- Bug bounty reports from major platforms
- Security conference presentations and whitepapers
- Analysis of PDF rendering library source code

EXCEL OUTPUT FEATURES:
- Comprehensive payload data with metadata
- CVE reference integration
- Browser-specific analysis sheets
- Category-wise payload breakdown
- Risk level assessment
- Advanced filtering and sorting capabilities
- Professional formatting for security research

Author: SNGWN
Legal Notice: For authorized security testing only.
"""

import json
import pandas as pd
import os
import sys
from datetime import datetime
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def find_latest_payload_database():
    """Find the most recent and comprehensive payload database file"""
    print("🔍 SEARCHING FOR PAYLOAD DATABASE FILES")
    print("=" * 45)
    
    # Search patterns for payload database files
    search_patterns = [
        'payload_database.json',  # Primary database file
        'merged_payload_database_*.json',
        'sophisticated_payload_database_*.json',
        'PDF/sophisticated_payload_database_*.json'
    ]
    
    found_files = []
    for pattern in search_patterns:
        files = glob.glob(pattern)
        found_files.extend(files)
    
    if not found_files:
        print("❌ No payload database files found")
        return None
    
    # Analyze files to find the most comprehensive one
    best_file = None
    max_payloads = 0
    
    print(f"📁 Found {len(found_files)} database files:")
    
    for file_path in found_files:
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
                payload_count = len(data.get('payloads', []))
                file_size = os.path.getsize(file_path)
                
                print(f"   📄 {os.path.basename(file_path)}")
                print(f"      Payloads: {payload_count}, Size: {file_size:,} bytes")
                
                if payload_count > max_payloads:
                    max_payloads = payload_count
                    best_file = file_path
                    
        except Exception as e:
            print(f"   ❌ Error reading {file_path}: {e}")
    
    if best_file:
        print(f"\n✅ Selected: {os.path.basename(best_file)} ({max_payloads} payloads)")
        return best_file
    
    return None

def load_payload_database(file_path):
    """Load and validate payload database"""
    print(f"\n📖 LOADING PAYLOAD DATABASE")
    print("=" * 32)
    
    try:
        with open(file_path, 'r') as f:
            data = json.load(f)
        
        payloads = data.get('payloads', [])
        metadata = data.get('metadata', {})
        
        print(f"✅ Successfully loaded {len(payloads)} payloads")
        print(f"📊 Database metadata:")
        print(f"   Generated: {metadata.get('generated_at', 'Unknown')}")
        print(f"   Version: {metadata.get('generator_version', 'Unknown')}")
        
        if 'breakdown' in metadata:
            breakdown = metadata['breakdown']
            print(f"   Browsers: {breakdown.get('browsers', {})}")
            print(f"   Categories: {breakdown.get('categories', {})}")
            print(f"   Risk levels: {breakdown.get('risk_levels', {})}")
        
        return data
        
    except Exception as e:
        print(f"❌ Error loading database: {e}")
        return None

def create_excel_workbook(data):
    """Create comprehensive Excel workbook with multiple sheets"""
    print(f"\n📊 CREATING EXCEL WORKBOOK")
    print("=" * 29)
    
    payloads = data.get('payloads', [])
    metadata = data.get('metadata', {})
    
    # Create main DataFrame from payloads
    df = pd.DataFrame(payloads)
    
    print(f"✅ Processing {len(df)} payloads for Excel export")
    
    # Create workbook with multiple sheets
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # 1. Main payload data sheet
    create_main_sheet(wb, df, metadata)
    
    # 2. Browser-specific sheets
    create_browser_sheets(wb, df)
    
    # 3. Category analysis sheet
    create_category_sheet(wb, df)
    
    # 4. CVE reference sheet
    create_cve_sheet(wb, df)
    
    # 5. Research summary sheet
    create_research_summary_sheet(wb, data)
    
    return wb

def create_main_sheet(wb, df, metadata):
    """Create main comprehensive payload sheet"""
    print("📋 Creating main payload sheet...")
    
    ws = wb.create_sheet("All Payloads", 0)
    
    # Add header information
    ws['A1'] = "XSS-PDF Payload Database - Comprehensive Security Research"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws['A3'] = f"Original Database: {metadata.get('generated_at', 'Unknown')}"
    ws['A4'] = f"Total Payloads: {len(df)}"
    ws['A5'] = "LEGAL NOTICE: For authorized security testing only"
    ws['A5'].font = Font(color="FF0000", bold=True)
    
    # Add payload data starting from row 7
    start_row = 7
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    header_row = start_row
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def create_browser_sheets(wb, df):
    """Create browser-specific analysis sheets"""
    print("🌐 Creating browser-specific sheets...")
    
    browsers = df['browser'].unique()
    
    for browser in browsers:
        browser_data = df[df['browser'] == browser]
        ws = wb.create_sheet(f"{browser.title()} Payloads")
        
        # Add header
        ws['A1'] = f"{browser.upper()} PDF Library - Targeted Payloads"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Total {browser} payloads: {len(browser_data)}"
        
        # Add browser-specific research info
        browser_info = {
            'chrome': 'PDFium - V8 engine exploitation, IPC abuse, process injection',
            'firefox': 'PDF.js - CSP bypass, SpiderMonkey exploitation, Content Security Policy evasion',
            'safari': 'PDFKit - macOS integration, WebKit messageHandlers, Objective-C bridge abuse',
            'adobe': 'Acrobat/Reader - Full JavaScript API, privilege escalation, file system access',
            'edge': 'Edge PDF - Windows integration, WebView exploitation, registry manipulation'
        }
        
        ws['A3'] = f"Focus: {browser_info.get(browser, 'Browser-specific exploitation techniques')}"
        
        # Add data starting from row 5
        start_row = 5
        for r in dataframe_to_rows(browser_data, index=False, header=True):
            ws.append(r)
        
        # Format header
        for col in range(1, len(browser_data.columns) + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

def create_category_sheet(wb, df):
    """Create payload category analysis sheet"""
    print("📂 Creating category analysis sheet...")
    
    ws = wb.create_sheet("Category Analysis")
    
    # Header
    ws['A1'] = "Payload Category Breakdown"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Category statistics
    category_counts = df['category'].value_counts()
    
    row = 3
    ws['A3'] = "Category"
    ws['B3'] = "Count"
    ws['C3'] = "Description"
    
    # Format header
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws[f'{col}3'].font = Font(color="FFFFFF", bold=True)
    
    category_descriptions = {
        'dom_access': 'Browser DOM manipulation from PDF context',
        'file_system': 'Local file system access and directory traversal',
        'command_execution': 'System command execution and process spawning',
        'sandbox_escape': 'PDF sandbox restriction bypasses',
        'network_exfiltration': 'Data exfiltration and covert channels'
    }
    
    row = 4
    for category, count in category_counts.items():
        ws[f'A{row}'] = category
        ws[f'B{row}'] = count
        ws[f'C{row}'] = category_descriptions.get(category, 'Security testing payload')
        row += 1
    
    # Auto-adjust columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 60

def create_cve_sheet(wb, df):
    """Create CVE reference analysis sheet"""
    print("🔒 Creating CVE reference sheet...")
    
    ws = wb.create_sheet("CVE References")
    
    # Header
    ws['A1'] = "CVE Security References - Research Foundation"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = "50+ CVE references across all PDF rendering libraries"
    
    # Extract all CVE references from payloads
    all_cves = set()
    for _, payload in df.iterrows():
        cve_ref = payload.get('cve_reference', '')
        if cve_ref:
            cves = [cve.strip() for cve in cve_ref.split(',')]
            all_cves.update(cves)
    
    # Sort CVEs
    sorted_cves = sorted(list(all_cves))
    
    row = 4
    ws['A4'] = "CVE ID"
    ws['B4'] = "Affected Component"
    ws['C4'] = "Associated Payloads"
    
    # Format header
    for col in ['A', 'B', 'C']:
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        ws[f'{col}4'].font = Font(color="FFFFFF", bold=True)
    
    row = 5
    for cve in sorted_cves:
        if cve.startswith('CVE-'):
            ws[f'A{row}'] = cve
            
            # Determine component based on CVE
            if any(x in cve for x in ['2019-5786', '2020-6418', '2021-21166']):
                component = "Chrome PDFium"
            elif any(x in cve for x in ['2020-12388', '2021-23953']):
                component = "Firefox PDF.js"
            elif any(x in cve for x in ['2019-8761', '2020-9715']):
                component = "Safari PDFKit"
            elif any(x in cve for x in ['2018-4996', '2019-7815']):
                component = "Adobe Reader"
            else:
                component = "Multi-platform"
            
            ws[f'B{row}'] = component
            
            # Count associated payloads
            payload_count = sum(1 for _, payload in df.iterrows() if cve in payload.get('cve_reference', ''))
            ws[f'C{row}'] = payload_count
            row += 1
    
    # Auto-adjust columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

def create_research_summary_sheet(wb, data):
    """Create research methodology and sources summary"""
    print("📚 Creating research summary sheet...")
    
    ws = wb.create_sheet("Research Summary")
    
    # Header
    ws['A1'] = "XSS-PDF Research Methodology & Sources"
    ws['A1'].font = Font(size=16, bold=True)
    
    research_content = [
        "",
        "RESEARCH FOUNDATION:",
        "=" * 20,
        "",
        "📄 ACADEMIC SOURCES:",
        "• PDF security research papers from security conferences",
        "• Analysis of PDF rendering library source code",
        "• Sandbox escape methodology studies",
        "• Cross-site scripting in PDF context research",
        "",
        "🔒 SECURITY REFERENCES:",
        "• 50+ CVE references across all PDF rendering libraries",
        "• Bug bounty reports from major platforms (HackerOne, Bugcrowd)",
        "• Security advisory disclosures",
        "• Vulnerability assessment methodologies",
        "",
        "🌐 BROWSER ANALYSIS:",
        "• Chrome (PDFium): V8 engine exploitation techniques",
        "• Firefox (PDF.js): Content Security Policy bypass methods",
        "• Safari (PDFKit): macOS integration exploit vectors",
        "• Adobe Reader: Full JavaScript API exploitation",
        "• Edge PDF: Windows integration security gaps",
        "",
        "🎯 PAYLOAD CATEGORIES:",
        "• DOM Access: Browser DOM manipulation from PDF context",
        "• File System: Local file system access and directory traversal",
        "• Command Execution: System command execution and process spawning",
        "• Sandbox Escape: PDF sandbox restriction bypasses",
        "• Network Exfiltration: Data exfiltration and covert channels",
        "",
        "📊 STATISTICAL BREAKDOWN:",
        f"• Total Unique Payloads: {len(data.get('payloads', []))}",
        f"• Research Sources: 50+ CVE references",
        f"• Browser Targets: 5 major PDF rendering engines",
        f"• Attack Categories: 5 distinct payload types",
        "",
        "⚖️ LEGAL & ETHICAL FRAMEWORK:",
        "• Designed for authorized security testing only",
        "• Educational and research purposes",
        "• Responsible disclosure practices",
        "• Compliance with applicable laws and regulations",
        "",
        "🛡️ DEFENSIVE APPLICATIONS:",
        "• PDF security assessment and hardening",
        "• Security awareness training materials",
        "• Penetration testing and red team exercises",
        "• Security control validation and testing",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}",
        "Tool: XSS-PDF Excel Exporter v1.0",
        "Author: SNGWN"
    ]
    
    for i, line in enumerate(research_content, 1):
        cell = ws[f'A{i}']
        cell.value = line
        if line.startswith(('📄', '🔒', '🌐', '🎯', '📊', '⚖️', '🛡️')):
            cell.font = Font(bold=True, color="2F5597")
        elif line and line[0] in "•":
            cell.font = Font(color="4472C4")
        elif "=" in line:
            cell.font = Font(bold=True)
    
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 80

def export_to_excel(output_file, data):
    """Export data to Excel with professional formatting"""
    print(f"\n💾 EXPORTING TO EXCEL")
    print("=" * 23)
    
    try:
        wb = create_excel_workbook(data)
        wb.save(output_file)
        
        file_size = os.path.getsize(output_file)
        print(f"✅ Excel export successful!")
        print(f"📄 File: {output_file}")
        print(f"📊 Size: {file_size:,} bytes")
        print(f"📋 Sheets: {len(wb.worksheets)}")
        
        # List all sheets
        print(f"📂 Sheet contents:")
        for sheet in wb.worksheets:
            print(f"   • {sheet.title}")
        
        return True
        
    except Exception as e:
        print(f"❌ Excel export failed: {e}")
        return False

def main():
    """Main execution function"""
    print("📊 XSS-PDF PAYLOAD DATABASE - EXCEL EXPORTER")
    print("=" * 48)
    print("Objective: Export comprehensive PDF security research data to Excel")
    print("Research Level: 50+ CVEs, academic papers, bug bounty reports")
    print("Legal: For authorized security testing only")
    print()
    
    # Find and load payload database
    db_file = find_latest_payload_database()
    if not db_file:
        print("❌ No payload database found. Please run the main script first.")
        return False
    
    data = load_payload_database(db_file)
    if not data:
        print("❌ Failed to load payload database.")
        return False
    
    # Generate output filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"xss_pdf_payload_database_{timestamp}.xlsx"
    
    # Export to Excel
    success = export_to_excel(output_file, data)
    
    if success:
        print(f"\n🎉 EXPORT COMPLETE")
        print("=" * 17)
        print(f"📊 Excel file ready for security research and testing")
        print(f"🔍 Professional formatting with multiple analysis sheets")
        print(f"📋 Same research data level as original PDF tools")
        print(f"⚖️ Remember: Use only for authorized security testing")
        return True
    else:
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)